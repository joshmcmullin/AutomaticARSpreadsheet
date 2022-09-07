import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


print("Initializing process.")
wb = openpyxl.load_workbook(filename="C:/Users/User/Desktop/Time Tracking Report.xlsx")
ws1 = wb.active

# Delete unwanted columns & data
print("Deleting unwanted columns.")
if ws1["N1"].value == "Service Item":
    ws1.delete_cols(idx = 14)
else:
    print("Error: Excel column N is not Service Item column.")


if ws1["L1"].value == "Billable Amount":
    ws1.delete_cols(idx = 12)
else:
    print("Error: Excel column L is not Billable Amount column.")


if ws1["K1"].value == "Billing Rate":
    ws1.delete_cols(idx = 11)
else:
    print("Error: Excel column K is not Billing Rate column.")


if ws1["J1"].value == "Billable Hours":
    ws1.delete_cols(idx = 10)
else:
    print("Error: Excel column J is not Billable Hours column.")


if ws1["H1"].value == "Project Budgeted Hours":
    ws1.delete_cols(idx = 8)
else:
    print("Error: Excel column H is not Project Budgeted Hours column.")


if ws1["G1"].value == "Task Budgeted Hours":
    ws1.delete_cols(idx = 7)
else:
    print("Error: Excel column G is not Task Budgeted Hours column.")


if ws1["F1"].value == "Task Title":
    ws1.delete_cols(idx = 6)
else:
    print("Error: Excel column F is not Task Title column.")


if ws1["E1"].value == "Accounting Period":
    ws1.delete_cols(idx = 5)
else:
    print("Error: Excel column E is not Accounting Period column.")

if ws1["B1"].value == "Role":
    ws1.delete_cols(idx = 2)
else:
    print("Error: Excel column B is not Role column.")

# Create a copy of the worksheet to be unedited. Create new blank sheets
print("Creating new worksheets.")
ws1.title = "Sheet 1"
ws2 = wb.copy_worksheet(wb["Sheet 1"])
ws2.title = "Sheet 2"
wsAdmin = wb.create_sheet("Admin")
wsTax = wb.create_sheet("Tax")
wsConcierge = wb.create_sheet("Concierge")
wsResolution = wb.create_sheet("Resolution")
wsSteveTime = wb.create_sheet("Steve Time")
wsDoNotBill = wb.create_sheet("Do not bill")


# Create headers, bold, & freeze
print("Creating headers.")
for ws in wb:
    ws['A1'].value = "Name"
    ws['B1'].value = "Client Name"
    ws['C1'].value = "Project Title"
    ws['D1'].value = "Actual Hours"
    ws['E1'].value = "Date"
    ws['F1'].value = "Notes"
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws.freeze_panes = ws['A2']


# Parse all rows to evaluate Admin Time or client meetings. If Admin time or client meetings, put in Admin sheet.
print("Moving Admin & Client Meeting tasks.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if cell.value == "Admin Time" or cell.value == "Client Meetings":
            wsAdmin.append(cell.value for cell in row)

# Delete Admin & Client Meeting rows from Sheet 2.
for row in ws2.iter_rows(min_row=2, max_row = ws1.max_row, min_col=1, max_col=6):
    for cell in row:
        if cell.value == "Admin Time" or cell.value == "Client Meetings":
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# Parse all rows to evaluate Tax Returns. If Tax return, put in Tax spreadsheet.
print("Moving tax return tasks.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if "Tax Return" in str(cell.value):
            wsTax.append(cell.value for cell in row)

# Delete Tax return rows from Sheet 2.
for row in ws2.iter_rows(min_row=2, max_row = ws1.max_row, min_col=1, max_col=6):
    for cell in row:
        if "Tax Return" in str(cell.value):
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# List of Concierge Clients to check Sheet 2 for.
conciergeClients = [] #conciergeClients have been redacted for privacy reasons.

# Parse Sheet 2 looking for Concierge Clients. Add Concierge Client rows to Concierge sheet.
print("Moving Concierge Client tasks.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if any(client in str(cell.value) for client in conciergeClients):
            wsConcierge.append(cell.value for cell in row)

# Remove Concierge Client data from Sheet 2.
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if any(client in str(cell.value) for client in conciergeClients):
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# Parse all rows to evaluate resolution projects. If a resolution project, put in Resolution sheet.
print("Moving resolution tasks.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if "Resolution" in str(cell.value):
            wsResolution.append(cell.value for cell in row)

# Delete Tax return rows from Sheet 2.
for row in ws2.iter_rows(min_row=2, max_row = ws1.max_row, min_col=1, max_col=6):
    for cell in row:
        if "Resolution" in str(cell.value):
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# Parse all rows to evaluate Steve's projects. If oen of Steve's projects, put in Steve Time sheet.
print("Moving Steve's projects.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if "Steve" in str(cell.value):
            wsSteveTime.append(cell.value for cell in row)

# Delete Steve's project rows from Sheet 2.
for row in ws2.iter_rows(min_row=2, max_row = ws1.max_row, min_col=1, max_col=6):
    for cell in row:
        if "Steve" in str(cell.value):
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# Parse all rows to evaluate unbillable entries. If found, put them in the Do Not Bill sheet.
print("Moving unbillable projects.")
for row in ws2.iter_rows(min_row = 2, max_row = ws1.max_row, min_col = 1, max_col = 6):
    for cell in row:
        if "do not bill" in str(cell.value).lower() or "don't bill" in str(cell.value).lower():
            wsDoNotBill.append(cell.value for cell in row)
            # for row_cell in ws2[cell.row]:
            #     row_cell.fill = PatternFill("solid", start_color="5cb800")  ### This is here incase I want to fill and not delete

# Delete unbillable project rows from Sheet 2.
for row in ws2.iter_rows(min_row=2, max_row = ws1.max_row, min_col=1, max_col=6):
    for cell in row:
        if "do not bill" in str(cell.value).lower() or "don't bill" in str(cell.value).lower():
            ws2["A" + str(cell.row)] = ""
            ws2["B" + str(cell.row)] = ""
            ws2["C" + str(cell.row)] = ""
            ws2["D" + str(cell.row)] = ""
            ws2["E" + str(cell.row)] = ""
            ws2["F" + str(cell.row)] = ""


# Remove empty rows from Sheet 2.
print("Cleaning up workbook. Please wait...")
i = 1
while i <= ws2.max_row:
    if not ws2["A" + str(i)].value:
        ws2.delete_rows(idx=i)
        i -= 1
    else:
        i += 1

wsAdmin.delete_rows(idx=2)
wsTax.delete_rows(idx=2)
wsConcierge.delete_rows(idx=2)
wsResolution.delete_rows(idx=2)
wsSteveTime.delete_rows(idx=2)
wsDoNotBill.delete_rows(idx=2)

editedMaxRow = str(ws2.max_row - 1 + wsAdmin.max_row - 1 + wsTax.max_row - 1 + wsConcierge.max_row - 1 +
                   wsResolution.max_row - 1 + wsSteveTime.max_row - 1 + wsDoNotBill.max_row - 1)

print("Process is complete. Saving workbook.")
print("Data in:")
print("Sheet 1: " + str(ws1.max_row - 1))
print("Sheet 2: " + str(ws2.max_row))
print("Admin: " + str(wsAdmin.max_row))
print("Tax: " + str(wsTax.max_row))
print("Concierge: " + str(wsConcierge.max_row))
print("Resolution: " + str(wsResolution.max_row))
print("Steve Time: " + str(wsSteveTime.max_row))
print("Do Not Bill: " + str(wsDoNotBill.max_row))
print("Total data in edited sheets: " + editedMaxRow)
wb.save("C:/Users/User/Desktop/Time Tracking Report - Adjusted.xlsx")

print("Press enter to close.")
input()